import datetime
import numpy as np
import openpyxl
import pandas as pd

from src.config import RAW_INTEGRATED_PM25_DIR, INTEGRATED_PM25_DIR
from src.data.archive_structure_parser import get_unzipped_directory_for_year
from src.data.file_operation import ensure_directory_exists
from src.data.index_query import get_all_sites, get_metadata
from src.data.text_transforms import rename_columns
from src.utils.logger_config import setup_logger

logger = setup_logger('data.extract_post_2010_data', 'extract_data.log')

# provided worksheet names
analyte_types = {'NT': 'Metals_ICPMS (Near-Total)', 'WS': 'Metals_ICPMS (Water-Soluble)'}

column_names_PM25 = [
    'NAPS Site ID', 'Sampling Date', 'Sample Type', 
    'PM2.5', 'PM2.5-MDL', 'PM2.5-Vflag', 
    'Pres.', 'Temp.', 'Start Time', 'End Time', 'Actual Volume'    
]


def get_file_path_post_2010(year, site_id):
    """
    Return a file path to extract data, which is different depending on the year.
    - inputs: 
        - year: year of the data (int)
        - site_id: NAPS site ID (int)
    - output: file_path (string)
    """
    unzipped_dir = get_unzipped_directory_for_year(year)
    # after 2016, files have a suffix '_EN'
    file_name = 'S' + str(site_id) + '_PM25_' + str(year) + ('.xlsx' if year < 2016 else '_EN.xlsx')
    
    file_path = str(RAW_INTEGRATED_PM25_DIR) + '/' + unzipped_dir + '/' + file_name
    return file_path


def find_header_row(sheet):
    """
    Return a row index of a header from a given sheet
    - input: worksheet of an xlsx file
    - row_index: row index (int) of the header of the data
    """
    row_idx = 0
    for col in sheet.columns:
        for cell in col:
            
            # the first found row which contains the value, 'NAPS Site ID', is the header row
            if cell.value == 'NAPS Site ID':
                break
                
            row_idx = row_idx + 1
            
        break  # should finish with only one loop
    return row_idx


def extract_PM25_vals(sheet, analyte_type):
    """
    Extract PM2.5 data from a sampler which is addressed based on the metal type.
    (Near Total metals are measured with Sampler #1 and 
    Water-soluble metals are measured with Sampler #2.)
    - input: 
        - sheet: worksheet of an xlsx file
        - analyte_type: 'NT' for Near total or 'WS' for Water-sluble data (string)
    - output: df: a DataFrame containing PM2.5 data
    """
    
    header_row = find_header_row(sheet)

    sampler = ('S-1' if analyte_type == 'NT' else 'S-2')
    sampler_to_mask = ('S-2' if analyte_type == 'NT' else 'S-1')
    
    cols = []
    for col in sheet.iter_cols():
        cells = []
        
        # Check the value of the top cell of a column 
        # (extract the columns with data other than sampler_to_mask)
        # AND the header row has a value (otherwise, an extra column will be extracted)
        if np.logical_not(col[0].value == sampler_to_mask) & np.logical_not(col[header_row].value is None):
            for cell in col:
                cells.append(cell.value)
            cols.append(cells)
    
    df = pd.DataFrame(cols, index=column_names_PM25).T.tail(-1).reset_index(drop=True)
    df.drop(range(0, header_row), inplace=True)
    
    df['sampler'] = sampler
    df.reset_index(inplace=True, drop=True)
    return df


def extract_metal_vals(sheet, analyte_type):
    rows = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        row_vals = []
        for cell in row:
            row_vals.append(cell.value)
        rows.append(row_vals)
        
    df = pd.DataFrame(rows)
    sampler = df.loc[0, 3]
    
    header_idx = df[df[0] == 'NAPS Site ID'].index.values[0]
    df.columns = df.iloc[header_idx]
    df = df.iloc[header_idx + 1:, :].reset_index(drop = True)
    df.columns.name = None
    
    df['sampler'] = sampler
    df['analyte_type'] = analyte_type   
    return df


def extract_ion_2010(sheet):
    '''Extract measured ion data from the file in or after 2010'''
    rows = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        row_vals = []
        for cell in row:
            row_vals.append(cell.value)
        rows.append(row_vals)
        
    df = pd.DataFrame(rows)
    
    # Sampler is stored at this cell so far
    # *** check if it is unchanged for the latest released datasets and update if needed ***
    sampler = df.loc[0, 3]
    
    header_idx = df[df[0] == 'NAPS Site ID'].index.values[0]
    df.columns = df.iloc[header_idx]
    df = df.iloc[header_idx + 1:, :].reset_index(drop = True)
    df.columns.name = None
    
    df['analyte_type'] = 'total'
    df['sampler'] = sampler
    return df


def extract_file(file_path, meta_df):
    """
    Extract ICP-MS measured data (metal and PM2.5) and IC measured data (ions).
    - inputs:
        - file_path:
        - meta_df:
    - outputs:
        - icpms_df: a DataFrame containing extracted ICP-MS measured data (metals and PM2.5)
        - ic_df: a DataFrame containing extracted IC measured data (ions)
    """
    book = openpyxl.load_workbook(file_path)
    
    # extract metal data and PM2.5 data and combine them
    icpms_df = pd.DataFrame()
    
    # if both NT and WS data exist, this loop is ran twice; otherwise one time
    for index, row in meta_df.iterrows():
        
        pm25_df = extract_PM25_vals(book['PM2.5'], row['analyte_type'])
        pm25_df = rename_columns(pm25_df)
        metal_df = extract_metal_vals(book[analyte_types[
                                      row['analyte_type']]], row['analyte_type'])
        metal_df = rename_columns(metal_df)
        
        merged_df = pm25_df.merge(metal_df, on=['site_id', 'sampling_date', 'sampling_type', 'sampler'])
        icpms_df = pd.concat([icpms_df, merged_df], ignore_index=True)
        
    # extract ion data even if ICPMS measured data does not exsit
    ic_df = pd.DataFrame()
    if ('Ions-Spec_IC' in book.sheetnames):
        ic_df = extract_ion_2010(book['Ions-Spec_IC'])
        ic_df = rename_columns(ic_df)
    
    return icpms_df, ic_df


def extract_post_2010():
    
    for year in list(range(2010, 2020)):
        
        logger.info(f'Start extracting PM2.5-Speciation data of {year}')
        
        # unique site list for the year
        site_ids = get_all_sites(year=year)
        for site_id in site_ids:
            
            # get the metainfo of the site to check if NT and/or WS data exists
            filtered_index_df = get_metadata(site_ids=[site_id], years=[year], instrument='ICPMS')
            analyte_type_index_data = filtered_index_df[['year', 'site_id', 'analyte_type']].drop_duplicates()        
            
            file_path = get_file_path_post_2010(year, site_id)
            metal_df, ion_df = extract_file(file_path, analyte_type_index_data)
            
            logger.debug(f'\t{ file_path[file_path.rindex("/") + 1:] }')
    
            if (len(metal_df) > 0):
                metal_df.to_csv(str(INTEGRATED_PM25_DIR) + '/' + str(year) + '_' + str(site_id) + '.csv', index = False)
    
            if (len(ion_df) > 0):
                ion_df.to_csv(str(INTEGRATED_PM25_DIR) + '/' + str(year) + '_' + str(site_id) + '_IC.csv', index = False)
            
        logger.info(f'Completed extracting data of {year}')
