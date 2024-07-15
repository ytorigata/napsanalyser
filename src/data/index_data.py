import datetime
import numpy as np
import openpyxl
import os
import pandas as pd
from pathlib import Path
import xlrd
from src.config import DATA_URLS_FILE, INDEX_CSV, RAW_INTEGRATED_PM25_DIR, \
STATIONS_RAW_CSV, STATIONS_CSV, METADATA_DIR
from src.data.archive_structure_parser import get_unzipped_directory_for_year
from src.data.file_operation import ensure_directory_exists
from src.data.text_transforms import remove_parentheses
from src.utils.logger_config import setup_logger

logger = setup_logger('data.index_data', 'index_data.log')

def extract_stations():
    """
    Extract station metadata from STATIONS_RAW_CSV to STATIONS_CSV and store it 
    into METADATA_DIR
    """

    # read CSV while skipping the rows with index of 0, 1, 2, 3, 5, and >788.
    # *** CHECK the row index and update the condition 
    # when StationsNAPS-StationsSNPA.csv is updated by the ECCC (every fall) ***
    stations = pd.read_csv(STATIONS_RAW_CSV, 
                           skiprows=(lambda x: (x < 4) | (x == 5) | (x > 790)), encoding='utf-8')
    
    # extract columns
    stations = stations[[
        'NAPS_ID', 'Station_Name', 'Status', 'City', 
        'Latitude', 'Longitude', 'Elevation', 'Start_Year', 'End_Year', 'Combined_Stations', 
        'SO2', 'CO', 'NO2', 'NO', 'NOX', 'O3', 'PM_25_Continuous', 'PM_10_Continuous', 
        'PM_2.5_RM', 'PM10-2.5', 'PM2.5_Speciation', 'VOC', 'Carbonyl', 'PAH', 
        'Site_Type', 'Urbanization', 'Neighbourhood', 'Land_Use', 'Scale'
    ]]
    
    # save the data as a meta-data of the measuring sites
    stations.rename(columns={
        'NAPS_ID': 'site_id', 'Station_Name': 'station_name', 
        'Site_Type': 'site_type', 'Land_Use': 'land_use'}, inplace=True)

    ensure_directory_exists(METADATA_DIR)
    stations.to_csv(STATIONS_CSV, index=False, encoding='utf-8')
    
    # check if the all sites were extracted: from 10101 to 129602 (for November 2023 data)
    print('The first three stations: ')
    display(stations.head(3))
    print('The last three stations: ')
    display(stations.tail(3))

# worksheet names for the datafile >= 2010
analyte_types = {
    'NT': 'Metals_ICPMS (Near-Total)', 
    'WS': 'Metals_ICPMS (Water-Soluble)',
    'total': 'Ions-Spec_IC'}

measured_metals = np.array([
    'aluminum', 'antimony', 'arsenic', 'barium', 'beryllium', 
    'cadmium', 'calcium', 'cerium', 'chromium', 'cobalt', 
    'copper', 'iron', 'lanthanum', 'lead', 'manganese', 
    'molybdenum', 'nickel', 'palladium', 'phosphorus', 'platinum', 
    'selenium', 'silver', 'strontium', 'sulfur', 'thallium', 
    'tin', 'titanium', 'uranium', 'vanadium', 'zinc'
])

measured_ions = np.array([
    'acetate', 'ammonium', 'barium', 'bromide', 'calcium', 
    'chloride', 'fluoride', 'formate', 'lithium', 'magnesium', 
    'manganese', 'msa', 'nitrate', 'nitrite', 
    'oxalate', 'phosphate', 'potassium', 'propionate', 'sodium', 
    'strontium', 'sulphate'
])

def is_relevant_file(file_name, year):
    """
    Check if the file is our target. For data before 2010, the file should 
    end with 'ICPMS.XLS', 'WICPMS.XLS', or 'IC.XLS'. For data in and after 2010, 
    the file shuold end with '{year}.xlsx' or '{year}_EN.xlsx'.
    """
    if year < 2010:
        return np.logical_or(
            file_name.endswith('ICPMS.XLS'), file_name.endswith('IC.XLS'))
    else:
        return np.logical_or(
            file_name.endswith(str(year) + '.xlsx'), file_name.endswith(str(year) + '_EN.xlsx'))

def site_id(file_name):
    """
    Return a site ID from a file name.
    - input: file_name (string)
    - output: site ID (int)
    """
    return (file_name[1:])[:file_name.index('_') - 1]

def header_row_before_2010(sheet):
    """
    Return the row index contains column names from a data file before 2010.
    - inputs:
        - sheet: worksheet (xlrd.sheet)
    - output: row_index_of_headr: row index (int)
    """
    # initialize a variable to store the row index
    row_index_of_header = None

   # iterate through each row in the sheet to look for a header row
    for row_idx in range(sheet.nrows):
        # flags to check if both 'Date' and 'NAPS ID' are found in the row
        found_date = False
        found_naps_id = False
        
        # iterate through each cell in the current row
        for col_idx in range(sheet.ncols):
            cell = sheet.cell(row_idx, col_idx)
            # Check if the cell contains text
            if cell.ctype == xlrd.XL_CELL_TEXT:
                # Check if the text is 'Date' or 'NAPS ID' (case insensitive)
                cell_value_lower = cell.value.lower()
                if 'date' in cell_value_lower:
                    found_date = True
                if 'naps id' in cell_value_lower:
                    found_naps_id = True
        
        # check if both 'Date' and 'NAPS ID' were found in the row
        if found_date and found_naps_id:
            row_index_of_header = row_idx
            break
    
    if row_index_of_header is None:
        logger.error("No row contains both 'Date' and 'NAPS ID'")
    
    return row_index_of_header

def header_row_in_and_after_2010(ws):
    """
    Return the row index contains column names from a data file in and after 2010.
    - inputs:
        - ws: worksheet (openpyxl.worksheet.worksheet.Worksheet)
    - output: row_index_of_headr: row index (int)
    """
    # initialize a variable to store the row index
    row_index_of_header = None
    
    # iterate through each row in the worksheet
    for row in ws.iter_rows():
        
        # convert all cell values in the row to lowercase and join them as a single string
        # (this makes it easier to check for the presence of 
        # 'NAPS Site ID' and 'Sampling Date' regardless of case)
        row_values_lower = [str(cell.value).lower() for cell in row if cell.value is not None]
        
        # check if 'NAPS Site ID' and 'Samplling Date' are both in the row
        if ('naps site id' in row_values_lower) & ('sampling date' in row_values_lower):
            row_index_of_header = row[0].row
            break
    
    if row_index_of_header is None:
        logger.error("No row contains both 'Sampling Date' and 'NAPS Site ID'")
    
    return row_index_of_header

def analytes_before_2010(item, year, instrument):
    """
    Return an array of analytes which contain at least one measurement 
    in the data file before 2010.
    - inputs:
        - item: datafile (pathlib.PosixPath)
        - year: year of the data (int)
        - instrument: an instrument used to measurement (string). 'ICPMS' or 'IC'
    - output: analytes: a numpy array containing analytes' full names (string)
    """
    
    # open a book, specifying encoding to avoid an error
    book = xlrd.open_workbook(item, encoding_override="cp1251")
    sheet = book.sheet_by_index(0)

    # look up the index of the header row
    header_row_index = header_row_before_2010(sheet)
    
    # get the column names from the header row
    column_names = sheet.row_values(header_row_index)
    
    # list to hold names of columns with content
    columns_with_content = []

    # iterate through each column by index
    for column_index in range(sheet.ncols):
        has_content = False
        # check each cell in the column, starting from the row below the header
        for row_idx in range(header_row_index + 1, sheet.nrows):
            cell = sheet.cell(row_idx, column_index)
            if cell.ctype != xlrd.XL_CELL_EMPTY and cell.ctype != xlrd.XL_CELL_BLANK:
                has_content = True
                break
        # if the column has content, add its name to the list
        if has_content:
            columns_with_content.append(column_names[column_index])

    columns_with_content = np.array(columns_with_content)
    # remove strings containing '-MDL'
    filtered_arr = columns_with_content[~np.char.find(columns_with_content, '-MDL') >= 0]
    # convert the remaining strings to lowercase
    lowercase_filtered_arr = np.char.lower(filtered_arr)

    lowercase_filtered_arr = np.sort(lowercase_filtered_arr)
    
    # keep strings in the pre-defined metals or ions
    analytes = ''
    if instrument == 'ICPMS':
        analytes = lowercase_filtered_arr[np.isin(lowercase_filtered_arr, measured_metals)]
    elif instrument == 'IC':
        analytes = lowercase_filtered_arr[np.isin(lowercase_filtered_arr, measured_ions)]
    
    return analytes

def analytes_in_and_after_2010(ws, year, analyte_type):
    """
    Return an array of analytes which contain at least one measurement 
    in the data file in and after 2010.
    - inputs:
        - ws: worksheet (openpyxl.worksheet.worksheet.Worksheet)
        - year: year of the data (int)
        - analyte_type: 'NT' (Near total), 'WS' (water-soluble), or 'total' (string)
    - output: analytes: a numpy array containing analytes' full names (string)
    """

    header_row_index = header_row_in_and_after_2010(ws)
    header_row = ws[header_row_index]
    
    # extract column names from the header row
    column_names = [cell.value for cell in header_row]
    
    # a list for the names of columns that have at least one non-empty cell
    columns_with_content = []
    
    # iterate through each column in the header row
    # **enumerate starting from 1 to match Excel's indexing**
    for col_index, cell in enumerate(header_row, start=1):
        has_content = False
        for row in ws.iter_rows(
            min_row=header_row_index + 1, min_col=col_index, max_col=col_index, values_only=True):
            
            # check if the cell is not empty; row[0] is the first analyte in a turple
            if (row[0] is not None) & (row[0] != ''):
                has_content = True
                break
        if has_content:
            columns_with_content.append(cell.value)
    
    # Now columns_with_content contains the names of the columns that 
    # have at least one non-empty cell below the header
    columns_with_content = np.array(columns_with_content)
    
    # remove strings containing '-MDL' or 'Flag'
    filtered_arr = columns_with_content[~np.char.find(columns_with_content, '-MDL') >= 0]
    filtered_arr = filtered_arr[~np.char.find(filtered_arr, 'Flag') >= 0]

    # trim abbreviation and parenthesis; e.g. Alminium (Al) -> Alminium
    # vectorize remove_parentheses and apply it to each analyte in the numpy array
    vectorized_remove = np.vectorize(remove_parentheses)
    trimmed_analytes = vectorized_remove(filtered_arr)
    
    # convert the remaining strings to lowercase
    lowercase_filtered_arr = np.char.lower(trimmed_analytes)

    lowercase_filtered_arr = np.sort(lowercase_filtered_arr)
    
    # keep strings in the pre-defined metals or ions
    if (analyte_type == 'NT') | (analyte_type == 'WS'):
        analytes = lowercase_filtered_arr[np.isin(lowercase_filtered_arr, measured_metals)]
    elif analyte_type == 'total':
        analytes = lowercase_filtered_arr[np.isin(lowercase_filtered_arr, measured_ions)]
    
    return analytes

def freq_before_2010(item, year, instrument):
    """
    Return the measurement frequency from two consequtive dates in column A. 
    Use this for reference because this check is not very accurate for some data.
    - inputs:
        - item: datafile (pathlib.PosixPath)
        - year: year of the data (int)
        - instrument: an instrument used to measurement (string). 'ICPMS' or 'IC'
    - output:
        - freq: frequency of measuring (int) 
    """
    
    def date_as_datetime(sheet, row_i, col_i, year):
        """
        Return the datetime which is converted from the data in a worksheet.
        ***The position of the first row is +1 only for 2009.***
        - inputs:
            - sheet: a workssheet (xlrd.sheet.Sheet) in a Excel book
            - row_i: a row index (int) for the starting date
            - col_i: a column index (int) for the dates
            - year: the year of the data (int)
        - output:
            dd: a gap of the two consequtive dates (int)
        """
        date1 = sheet.cell_value(row_i if year < 2009 else row_i + 1, col_i)
        date1_as_datetime = datetime.datetime(*xlrd.xldate_as_tuple(date1, book.datemode))
        return date1_as_datetime
    
    # open a book, specifying encoding to avoid an error
    book = xlrd.open_workbook(item, encoding_override="cp1251")
    sheet = book.sheet_by_index(0)
    max_row = sheet.nrows
    
    # initialilze the output value with 100 (to manually modify for some datafile)
    freq = 100

    row_idxs = {}
    row_idxs_ICPMS = {
        'set1_start': 2, 'set1_end': 3,
        'set2_start': 3, 'set2_end': 4,
        'set3_start': 4, 'set3_end': 5
    }
    # In IC-measured data file, a regular measurement and field blank appears in turn,
    # so, skip a row to select "consecutive" date.
    row_idxs_IC = {
        'set1_start': 2, 'set1_end': 4,
        'set2_start': 4, 'set2_end': 6,
        'set3_start': 6, 'set3_end': 8
    }
    if instrument == 'ICPMS':
        row_idxs = row_idxs_ICPMS.copy()
    else:
        row_idxs = row_idxs_IC.copy()
    
    # check the first set of consecutive dates
    gap0 = (date_as_datetime(sheet, row_idxs['set1_end'], 0, year) - 
            date_as_datetime(sheet, row_idxs['set1_start'], 0, year)).days
    
    # check the second set of consecutive dates if the number of rows >= 5
    if max_row >= 5:
        gap1 = (date_as_datetime(sheet, row_idxs['set2_end'], 0, year) - 
                date_as_datetime(sheet, row_idxs['set2_start'], 0, year)).days
        
        # if the first gap and second gap agree, assign the value as a frequency
        freq = gap0 if (gap0 == gap1) else 100

    # check the third set of consecutive dates if the number of rows >= 6
    if max_row >= 6:
        gap2 = (date_as_datetime(sheet, row_idxs['set3_end'], 0, year) - 
                date_as_datetime(sheet, row_idxs['set3_start'], 0, year)).days
        
        if (gap1 == gap2):
            freq = gap1
        elif (gap0 == gap2):
            freq = gap2
    
    # if any two gaps do not agree, use the typical frequency if it was detected at least one time
    if (freq == 100) & ((gap0 == 3) | (gap0 == 6)):
        freq = gap0

    return freq

def freq_in_and_after_2010(sheet, year):
    '''
    Return the measurement frequency from two consequtive dates in column B.
    - input:
        - sheet: Excel worksheet (openpyxl.worksheet.worksheet.Worksheet)
        - year: tbe year of the data (int)
    - output: freq: a gap of the two consecutive dates (int). For error, set 100.
    '''
    
    # the row indexes are randomly picked up...depending on the number of row in a file
    # you may have to adjust these indexes for a datafile with fewer data
    d0 = sheet['B15'].value
    d1 = sheet['B16'].value

    freq = 100
    if (d0 is not None) & (d1 is not None):
        freq = (d1 - d0).days
    return freq

def create_row_before_2010(item, year):
    """
    Return an array of rows containing the metadata for a data file before 2010.
    - inputs:
        - item: datafile (pathlib.PosixPath)
        - year: year of the data
    - output:
        - rows: an array of rows of the metadata of a data file
    """
    file_name = item.name

    # determine analyte_type and instrument from the suffix of the file
    analyte_type = ''
    instrument = ''
    if file_name.endswith('_ICPMS.XLS'):
        analyte_type = 'NT'
        instrument = 'ICPMS'
    elif file_name.endswith('_WICPMS.XLS'):
        analyte_type = 'WS'
        instrument = 'ICPMS'
    elif file_name.endswith('_IC.XLS'):
        analyte_type = 'total'
        instrument = 'IC'
    else:
        logger.error(f'{file_name} is not expected neither for ICPMS nor IC measured data.')
    
    analytes = analytes_before_2010(item, year, instrument)
    frequency = freq_before_2010(item, year, instrument)
    
    rows = []
    for analyte in analytes:
        rows.append({
            'year': year,
            'site_id': int(site_id(file_name)),
            'analyte': analyte,
            'analyte_type': analyte_type,
            'instrument': instrument,
            'frequency': frequency
        })
    
    return rows

def create_row_in_and_after_2010(item, year):
    """
    Return an array of one or two rows which contains the metainfo 
    of the data in and after 2010.
    - inputs:
        - item: datafile (pathlib.PosixPath)
        - year: year of the data
    - output:
        - rows: an array of one or two rows of the metainfo of the data
    """
    file_name = item.name
    book = openpyxl.load_workbook(item)
    
    # check if the worksheet for NT and/or WS data exists
    rows = []
    for analyte_type, sheet_name in analyte_types.items():   
        if sheet_name in book.sheetnames:

            analytes = analytes_in_and_after_2010(book[sheet_name], year, analyte_type)
            
            instrument = ''
            if (analyte_type == 'NT') | (analyte_type == 'WS'):
                instrument = 'ICPMS'
            elif analyte_type == 'total':
                instrument = 'IC'
            
            frequency = freq_in_and_after_2010(book[sheet_name], year)
            
            for analyte in analytes:
                rows.append({
                    'year': year,
                    'site_id': int(site_id(file_name)),
                    'analyte': analyte,
                    'analyte_type': analyte_type,
                    'instrument': instrument,
                    'frequency': frequency
                })
    return rows


def index_dataset_attributes():
    """
    Create an index file INDEX_CSV to show the availability of integrated data  
    listed in DATA_URLS_FILE. Assume the raw data are stored in RAW_INTEGRATED_PM25_DIR
    """
    url_df = pd.read_csv(DATA_URLS_FILE)
    integrated_df = url_df[url_df['type'] == 'integrated_pm25'].copy()
    years = integrated_df.sort_values('year')['year'].squeeze().unique()
    
    # check the presence of the data of our interest (Near Total and Water-sluble speciation data)
    rows_list = []
    for year in years:
        
        logger.info(f'Start scanning the source directory of {year} >>>')
        
        # retrieve an unzipped directory for a particular year
        target_dir = Path(str(RAW_INTEGRATED_PM25_DIR) + '/' + get_unzipped_directory_for_year(year))
        
        for item in target_dir.iterdir():
            
            # check a file if it is relevant
            if is_relevant_file(item.name, year):
                row = []
                if year < 2010:
                    row.extend(create_row_before_2010(item, year))
                else:
                    row.extend(create_row_in_and_after_2010(item, year))
                if row is not None:
                    rows_list.extend(row)
    
        logger.info(f'<<< Complete scanning the data of {year}.')
    
    # save the metadata to a CSV file
    metadata_df = pd.DataFrame.from_records(rows_list)
    metadata_df.sort_values(['year', 'site_id', 'analyte'], inplace=True)
    metadata_df = metadata_df.reset_index(drop=True)
    metadata_df.to_csv(INDEX_CSV, index=False, encoding='utf-8')

def apply_manually_checked_frequency(index_df, CHECKED_FREQUENCY, INDEX_CSV):
    """
    Correct the index CSV by applying manually-checked frequency data.
    - inputs:
        - index_df: a DataFrame which is the lodaded index CSV
        - CHECKED_FREQUENCY: the file path to the CSV file containing correct frequencies
        - INDEX_CSV: a file path (string) to the index CSV
    - output: updated_df: a corrected DataFrame
    """
    
    # marge the metadata and eye-checked CSV on key columns
    checked_freq = pd.read_csv(CHECKED_FREQUENCY)
    df_merged = pd.merge(
        index_df, checked_freq, 
        on=['year', 'site_id', 'analyte_type', 'instrument'], how='left', suffixes=('', '_new'))
    
    # Update 'frequency' in index_df by using checked_freq where available
    df_merged['frequency'] = np.where(
        df_merged['frequency_new'].notna(), df_merged['frequency_new'], df_merged['frequency'])
    
    df_merged['frequency'] = df_merged['frequency'].astype(int)
    
    # drop the temporary '_new' column
    updated_df = df_merged.drop(columns=['frequency_new'])
    
    updated_df.to_csv(INDEX_CSV, index=False)
    return updated_df
    
def drop_entries_with_too_few_measurements(index_df, INDEX_CSV):
    """
    Drop the entries which contains too few measurements from a DataFrame of the index CSV.
    - inputs:
        - index_df: a DataFrame
        - INDEX_CSV: a file path (string) to the index CSV
    - output: index_df: the updated DataFrame
    """
    
    will_be_dropped = index_df[(
        (index_df['year'] == 2007) & 
        (index_df['site_id'] == 70301) & 
        (index_df['analyte_type'] == 'WS')) | (
            (index_df['year'] == 2016) & 
            (index_df['site_id'] == 129302) & 
            (index_df['analyte_type'] == 'total')) | (
            (index_df['year'] == 2017) & 
            (index_df['site_id'] == 60610) & 
            (index_df['analyte_type'] == 'total'))]
    
    logger.info(f'{len(will_be_dropped)} rows will be dropped from the index CSV with {len(index_df)} entries')
    
    index_df = index_df.drop(index_df[(
        (index_df['year'] == 2007) & 
        (index_df['site_id'] == 70301) & 
        (index_df['analyte_type'] == 'WS')) | (
            (index_df['year'] == 2016) & 
            (index_df['site_id'] == 129302) & 
            (index_df['analyte_type'] == 'total')) | (
            (index_df['year'] == 2017) & 
            (index_df['site_id'] == 60610) & 
            (index_df['analyte_type'] == 'total'))].index)
    
    index_df.to_csv(INDEX_CSV, index=False, encoding='utf-8')
    return index_df

def update_index_with_major_frequency(index_df, INDEX_CSV):
    """
    Update frequencies in a DataFrame for the index CSV 
    to the majority one when there are more than one. Save it to the file.
    - inputs:
        - index_df: a DataFrame
        - INDEX_CSV: a file path (string) to the index CSV
    - output: index_df: the updated DataFrame
    """
    index_df.loc[(index_df['year'] == 2006) & 
                (index_df['site_id'] == 100119) & 
                (index_df['analyte_type'] == 'WS'), ['frequency']] = 3
    index_df.loc[(index_df['year'] == 2014) & 
                (index_df['site_id'] == 129003) & 
                (index_df['analyte_type'] == 'NT'), ['frequency']] = 3
    index_df.loc[(index_df['year'] == 2015) & 
                (index_df['site_id'] == 129003) & 
                (index_df['analyte_type'] == 'NT'), ['frequency']] = 3
    
    index_df.to_csv(INDEX_CSV, index=False)
    return index_df
